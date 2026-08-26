from unittest.mock import patch
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from virtual_feed_mvp.core import (
    PDPExtractor,
    LGSitemapResolver,
    ProductInput,
    Evidence,
    fit_text,
    load_products_from_file,
    process_product,
    rule_body,
    parse_flattened_wide_table,
    parse_tabular_text,
    extract_product_codes,
    extract_msrp,
    product_title,
    classify_intent,
    category_from_url,
    exclusion_reason,
    is_complete_copy,
)
from virtual_feed_mvp.exporter import build_workbook


ROOT = Path(__file__).resolve().parents[1]


class CoreTests(unittest.TestCase):
    def test_wide_sample_becomes_four_groups_of_ten(self):
        products = load_products_from_file("sample.csv", (ROOT / "sample_uk_40.csv").read_bytes())
        self.assertEqual(40, len(products))
        counts = {category: sum(p.category == category for p in products) for category in {p.category for p in products}}
        self.assertEqual({"REF": 10, "W/M": 10, "LTV": 10, "MNT": 10}, counts)

    def test_flattened_clipboard_table_is_restored(self):
        products = load_products_from_file("sample.csv", (ROOT / "sample_uk_40.csv").read_bytes())
        lines = ["�", "Country", "#", "REF", "W/M", "LTV", "MNT", "영국"]
        for row_no in range(1, 11):
            lines.append(str(row_no))
            row = [p for p in products if (int(p.request_id[1:]) - 1) // 4 + 1 == row_no]
            lines.extend(p.sku for p in row)
        restored = parse_flattened_wide_table("\n".join(lines))
        self.assertEqual(40, len(restored))
        self.assertEqual({"REF": 10, "W/M": 10, "LTV": 10, "MNT": 10}, {c: sum(p.category == c for p in restored) for c in {p.category for p in restored}})

    def test_sitemap_selects_direct_category_match(self):
        urls = [
            "https://www.lg.com/uk/tncs/promo-gsxv91mcae/",
            "https://www.lg.com/uk/fridge-freezers/american-style-fridge-freezers/gsxv91mcae1/",
            "https://www.lg.com/uk/fridge-freezers/bundle/gsxv91mcae-ms2032/",
        ]
        selected = LGSitemapResolver.select_url("GSXV91MCAE", "REF", urls)
        self.assertTrue(selected.endswith("/gsxv91mcae1/"))

    def test_pim_features_generate_traceable_drafts(self):
        products = load_products_from_file("pim.csv", (ROOT / "sample_pim_export.csv").read_bytes())
        products[0].original_sku = "DEMO-SKU.AUK"
        result = process_product(products[0], generator_mode="rules")
        self.assertEqual(3, len(result.evidence))
        self.assertEqual(3, len(result.feeds))
        self.assertTrue(all(feed.evidence_ids for feed in result.feeds))
        self.assertTrue(all(feed.title_chars <= 30 and feed.body_chars <= 60 for feed in result.feeds))
        self.assertEqual(["DEMO-SKU.AUK_1", "DEMO-SKU.AUK_2", "DEMO-SKU.AUK_3"], [feed.virtual_sku for feed in result.feeds])

    def test_multiple_urls_accept_newlines_or_commas(self):
        urls = [
            "https://www.lg.com/uk/fridge-freezers/a/model-one/",
            "https://www.lg.com/uk/laundry/a/model-two/",
            "https://www.lg.com/uk/monitors/a/model-three/",
        ]
        self.assertEqual(3, len(parse_tabular_text("\n".join(urls))))
        self.assertEqual(3, len(parse_tabular_text(",".join(urls))))

    def test_sku_without_source_is_not_generated(self):
        result = process_product(ProductInput("R1", "SKU-ONLY", country="KR"), generator_mode="rules")
        self.assertFalse(result.feeds)
        self.assertIn("SOURCE_REQUIRED", {issue.code for issue in result.issues})

    def test_key_feature_extraction(self):
        html = """
        <html><body><h1>Demo SKU-1</h1><h2>Key Features</h2>
        <ul><li>Knock twice and see inside with InstaView</li>
        <li>Keep food fresher with LinearCooling</li></ul><div>More</div></body></html>
        """
        product = ProductInput("R1", "SKU-1", "REF", url="https://example.test/product")
        _, evidence, issues = PDPExtractor().extract(product, html)
        self.assertEqual(2, len(evidence))
        self.assertFalse(issues)

    def test_embedded_product_json_yields_market_sku(self):
        html = r'''<div data-pim-sku="GSXV91MCAE.AMCQLGU.EEUK.UK.C"></div>
        <script>{\"salesModelCode\":\"GSXV91MCAE\",\"salesSuffixCode\":\"AMCQLGU\"}</script>'''
        model, market_sku = extract_product_codes(html)
        self.assertEqual("GSXV91MCAE", model)
        self.assertEqual("GSXV91MCAE.AMCQLGU.EEUK.UK.C", market_sku)

    def test_extract_msrp_prefers_regular_price(self):
        html = '{"regularPrice":"249.98","price":"199.00"}'
        self.assertEqual("249.98", extract_msrp(html))

    def test_product_title_contains_brand_and_product_type(self):
        product = ProductInput(
            request_id="TV1", sku="50UA73006LA", category="LTV", country="UK",
            product_name='LG 50" UHD AI 4K Smart TV 2025',
        )
        evidence = Evidence(
            claim_id="TV-C01", sku=product.sku, category="LTV", source_type="PDP",
            source_record_id="url", source_section="Key Features",
            source_text="AI-optimized 4K picture quality", intent_candidate="Picture Quality",
            validation_status="Needs Review",
        )
        title = product_title(product, evidence, 30)
        self.assertTrue(title.startswith("LG "))
        self.assertIn("TV", title)
        self.assertLessEqual(len(title), 30)

    def test_fridge_title_uses_editable_intent_label(self):
        product = ProductInput(request_id="REF1", sku="GSXV91MCAE", category="REF")
        evidence = Evidence(
            claim_id="REF-C01", sku=product.sku, category="REF", source_type="PDP",
            source_record_id="url", source_section="Key Features", source_text="InstaView",
            intent_candidate="Easy Access", validation_status="Needs Review",
        )
        self.assertEqual("LG Fridge Freezer: Easy Access", product_title(product, evidence, 30))

    def test_long_internal_intent_uses_short_title_label(self):
        product = ProductInput(request_id="REF2", sku="GSXV91MCAE", category="REF")
        evidence = Evidence(
            claim_id="REF-C04", sku=product.sku, category="REF", source_type="PDP",
            source_record_id="url", source_section="Key Features", source_text="LG ThinQ",
            intent_candidate="Smart Convenience", validation_status="Needs Review",
        )
        title = product_title(product, evidence, 30)
        self.assertEqual("LG Fridge Freezer: Smart Care", title)
        self.assertLessEqual(len(title), 30)

    def test_category_is_inferred_from_resolved_pdp_before_title_mapping(self):
        product = ProductInput(request_id="R-REF", sku="GSXV91MCAE", country="UK", language="en-GB")
        url = "https://www.lg.com/uk/fridge-freezers/american-style-fridge-freezers/gsxv91mcae/"
        html = """
        <html><body><h1>LG GSXV91MCAE Fridge Freezer</h1>
        <script>{"modelName":"GSXV91MCAE","sku":"GSXV91MCAE.AMCQLGU.EEUK.UK.C"}</script>
        <h2>Key Features</h2><p>InstaView Door-in-Door for easy access</p><p>More</p>
        </body></html>
        """
        from unittest.mock import patch
        with patch.object(LGSitemapResolver, "resolve", return_value=url), patch.object(PDPExtractor, "fetch", return_value=html):
            result = process_product(product, generator_mode="rules")
        self.assertEqual("REF", result.product.category)
        self.assertEqual(1, len(result.feeds))
        self.assertEqual("LG Fridge Freezer: Easy Access", result.feeds[0].brand_title)

    def test_unmapped_intent_does_not_cancel_all_feeds(self):
        product = ProductInput(request_id="R-REF", sku="GSXV91MCAE", category="REF", country="UK")
        evidence = Evidence(
            claim_id="REF-C99", sku=product.sku, category="REF", source_type="PDP",
            source_record_id="url", source_section="Key Features", source_text="A general product feature",
            intent_candidate="Product Benefit", validation_status="Needs Review",
        )
        self.assertEqual("LG Fridge Freezer: Key Benefit", product_title(product, evidence, 30))

    def test_feature_rules_replace_generic_key_benefit(self):
        self.assertEqual("Auto Dosing", classify_intent(
            "Use the precise amount of detergent and softener each load", "W/M", "UK"
        ))
        self.assertEqual("Picture Quality", classify_intent(
            "Colour Fidelity for accurate lifelike colours", "LTV", "UK"
        ))

    def test_dimension_only_feature_is_excluded(self):
        self.assertEqual("SPEC_ONLY", exclusion_reason(
            "Dimension (mm): 913(W) x 1790(H) x 735(D)", "REF", "UK"
        ))

    def test_long_copy_is_rewritten_as_complete_clause(self):
        body = rule_body("LG BIG-In models can offer an ultra large 13kg capacity in a compact cabinet", 60)
        self.assertLessEqual(len(body), 60)
        self.assertTrue(is_complete_copy(body))
        self.assertNotIn("in a", body[-5:].lower())

    def test_ambiguous_generic_price_is_not_exported(self):
        self.assertEqual("", extract_msrp('{"price":"50.00","screenSize":"50"}'))

    def test_br_category_is_inferred_from_portuguese_url(self):
        self.assertEqual("REF", category_from_url("https://www.lg.com/br/geladeiras/geladeiras-inverse/gc-b569nqlc/"))

    def test_resolver_searches_when_sitemap_has_no_matching_pdp(self):
        url = "https://www.lg.com/br/geladeiras/geladeiras-inverse/gc-b569nqlc/"
        with patch.object(LGSitemapResolver, "_load_urls", return_value=[]), patch(
            "virtual_feed_mvp.core.search_lg_pdp", return_value=[url]
        ) as search:
            self.assertEqual(url, LGSitemapResolver.resolve("BR", "GC-B569NQLC", ""))
            search.assert_called_once_with("br", "GC-B569NQLC")

    def test_portuguese_feature_heading_is_extracted(self):
        html = """
        <html><body><h1>Geladeira LG GC-B569NQLC</h1>
        <script>{"modelName":"GC-B569NQLC","sku":"GC-B569NQLC.AMCFSBS"}</script>
        <h2>Principais recursos</h2>
        <p>Alta Eficiência Energética A</p><p>Smart Inverter Compressor</p>
        <p>Door Cooling+</p><p>Gaveta Fresh Zone</p><p>Mais</p>
        </body></html>
        """
        product = ProductInput("BR-1", "GC-B569NQLC", "REF", "BR", "pt-BR", url="https://www.lg.com/br/geladeiras/geladeiras-inverse/gc-b569nqlc/")
        _, evidence, _ = PDPExtractor().extract(product, html)
        self.assertGreaterEqual(len(evidence), 2)
        self.assertIn("Energy Saving", [e.intent_candidate for e in evidence])

    def test_metric_is_not_left_dangling_at_cutoff(self):
        source = "Ice and water dispenser with UVNano™ to eliminate 99.99% bacteria."
        fitted = rule_body(source, 60)
        self.assertLessEqual(len(fitted), 60)
        self.assertFalse(fitted.endswith("99.99%"))
        self.assertIn("bacteria", fitted)

    def test_workbook_has_delivery_feed_schema(self):
        products = load_products_from_file("pim.csv", (ROOT / "sample_pim_export.csv").read_bytes())
        result = process_product(products[0], generator_mode="rules")
        with tempfile.TemporaryDirectory() as tmp:
            path = build_workbook([result], Path(tmp) / "result.xlsx", {"test": True})
            wb = load_workbook(path, read_only=True, data_only=True)
            self.assertEqual(["Feeds"], wb.sheetnames)
            headers = [cell.value for cell in next(wb["Feeds"].iter_rows(max_row=1))]
            self.assertEqual(["id", "title", "link", "image_link", "description", "price"], headers)
            wb.close()

    def test_rule_body_does_not_invent_thinq_benefit(self):
        source = "LG ThinQ app enables remote control and alerts"
        self.assertEqual(rule_body(source, 60), source)

    def test_explicit_rule_mode_reports_correct_reason(self):
        product = ProductInput(
            request_id="R-RULE", sku="RULE-SKU", category="REF",
            country="UK", language="en-GB",
            key_features=["LG ThinQ app enables remote control and alerts"],
        )
        result = process_product(product, generator_mode="rules")
        finding = next(issue for issue in result.issues if issue.code == "RULE_DRAFT")
        self.assertIn("규칙 기반 모드를 선택", finding.message)
        self.assertNotIn("API 키가 없어", finding.message)


if __name__ == "__main__":
    unittest.main()
