from __future__ import annotations

import csv
import io
import json
import os
import re
import threading
import time
import urllib.error
import urllib.request
from urllib.parse import urljoin
from dataclasses import asdict, dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable

from .browser_fetch import dump_dom, search_lg_pdp


WIDE_CATEGORIES = {
    "ref": "REF",
    "w/m": "W/M",
    "wm": "W/M",
    "ltv": "LTV",
    "mnt": "MNT",
}


def load_country_registry() -> dict[str, Any]:
    path = Path(__file__).with_name("country_registry.json")
    return json.loads(path.read_text(encoding="utf-8"))


def enabled_country_sites() -> list[dict[str, Any]]:
    return [item for item in load_country_registry().get("sites", []) if item.get("enabled")]


def country_config(value: str) -> dict[str, Any] | None:
    needle = (value or "").strip().casefold()
    sites = load_country_registry().get("sites", [])
    # Prefer exact site key; market aliases only resolve when unambiguous/default.
    for item in sites:
        if needle == str(item.get("site_key", "")).casefold():
            return item
    candidates = []
    for item in sites:
        aliases = [item.get("market_code", ""), item.get("lg_site_code", ""), *item.get("aliases", [])]
        if needle in {str(alias).casefold() for alias in aliases}:
            candidates.append(item)
    if len(candidates) == 1:
        return candidates[0]
    return next((item for item in candidates if item.get("is_default")), None)


@dataclass
class ProductInput:
    request_id: str
    sku: str
    category: str = ""
    country: str = "UK"
    language: str = "en-GB"
    url: str = ""
    product_name: str = ""
    key_features: list[str] = field(default_factory=list)
    source_record_id: str = ""
    original_sku: str = ""
    image_link: str = ""
    price: str = ""


@dataclass
class Evidence:
    claim_id: str
    sku: str
    category: str
    source_type: str
    source_record_id: str
    source_section: str
    source_text: str
    intent_candidate: str
    validation_status: str
    notes: str = ""


@dataclass
class Feed:
    sku: str
    virtual_sku: str
    category: str
    country: str
    language: str
    feed_no: int
    intent: str
    brand_title: str
    title_chars: int
    brand_body_copy: str
    body_chars: int
    evidence_ids: str
    validation_status: str
    review_notes: str = ""


@dataclass
class Issue:
    sku: str
    stage: str
    severity: str
    code: str
    message: str


@dataclass
class ProductResult:
    product: ProductInput
    feeds: list[Feed] = field(default_factory=list)
    evidence: list[Evidence] = field(default_factory=list)
    issues: list[Issue] = field(default_factory=list)
    elapsed_seconds: float = 0.0


def normalize_sku(value: str) -> str:
    value = (value or "").strip().upper()
    value = value.split(".", 1)[0]
    return re.sub(r"[^A-Z0-9-]", "", value)


def normalize_original_sku(value: str) -> str:
    return re.sub(r"[^A-Z0-9.\-]", "", (value or "").strip().upper())


def category_from_url(url: str) -> str:
    path = url.lower()
    if "fridge" in path or "/geladeiras/" in path or "/refrigeradores/" in path:
        return "REF"
    if "/laundry/" in path or "/lavanderia/" in path or "/maquinas-de-lavar/" in path:
        return "W/M"
    if "/monitor" in path or "/monitores/" in path:
        return "MNT"
    if "/tv" in path or "/tvs/" in path or "television" in path or "/televisores/" in path:
        return "LTV"
    return ""


def _html_field(html: str, name: str) -> str:
    patterns = [
        rf'"{re.escape(name)}"\s*:\s*"([^"\\]+)"',
        rf'\\"{re.escape(name)}\\"\s*:\s*\\"([^"\\]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, html, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return ""


def extract_product_codes(html: str) -> tuple[str, str]:
    """Return (sales model, full product ID) from LG PDP HTML or embedded JSON."""
    sales_model = _html_field(html, "salesModelCode")
    sales_suffix = _html_field(html, "salesSuffixCode")
    sku_candidates = re.findall(r'data-pim-sku=["\']([^"\']+)', html, flags=re.IGNORECASE)
    embedded_sku = _html_field(html, "sku")
    if embedded_sku:
        sku_candidates.insert(0, embedded_sku)

    prefix = normalize_original_sku(f"{sales_model}.{sales_suffix}").rstrip(".")
    normalized_candidates = [normalize_original_sku(value) for value in sku_candidates]
    full_product_id = next(
        (value for value in normalized_candidates if value and (not prefix or value.startswith(prefix + "."))),
        normalized_candidates[0] if normalized_candidates else "",
    )
    if sales_model and full_product_id:
        return normalize_sku(sales_model), full_product_id
    if sales_model and sales_suffix:
        return normalize_sku(sales_model), prefix

    parts = full_product_id.split(".")
    if len(parts) >= 2:
        return normalize_sku(parts[0]), full_product_id
    return normalize_sku(parts[0] if parts else ""), ""


def _clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _is_url(value: str) -> bool:
    return value.lower().startswith(("https://", "http://"))


def _features_from_row(row: dict[str, Any]) -> list[str]:
    features: list[str] = []
    for key, value in row.items():
        header = _clean_header(key)
        text = str(value or "").strip()
        if not text:
            continue
        if header.startswith("key_feature") or header.startswith("key feature"):
            features.append(text)
        elif header in {"key_features", "key features", "features"}:
            features.extend(x.strip() for x in re.split(r"\s*\|\s*|\r?\n", text) if x.strip())
    return list(dict.fromkeys(features))


def rows_to_products(headers: list[str], rows: Iterable[list[Any]]) -> list[ProductInput]:
    cleaned = [_clean_header(h) for h in headers]
    wide_columns = [(idx, WIDE_CATEGORIES[h]) for idx, h in enumerate(cleaned) if h in WIDE_CATEGORIES]
    products: list[ProductInput] = []
    last_country = "UK"

    if wide_columns:
        country_idx = next((i for i, h in enumerate(cleaned) if h in {"country", "market", "국가"}), None)
        url_map: dict[str, str] = {}
        for row in rows:
            values = list(row) + [""] * max(0, len(headers) - len(row))
            if country_idx is not None and str(values[country_idx] or "").strip():
                raw_country = str(values[country_idx]).strip()
                last_country = "UK" if raw_country in {"영국", "United Kingdom", "GB"} else raw_country
            for idx, category in wide_columns:
                raw = str(values[idx] or "").strip()
                if not raw:
                    continue
                sku = normalize_sku(raw)
                if not sku:
                    continue
                products.append(ProductInput(
                    request_id=f"R{len(products)+1:03d}", sku=sku, category=category,
                    country=last_country or "UK", language="en-GB", url=url_map.get(sku, ""),
                    original_sku=normalize_original_sku(raw),
                ))
        return products

    header_map = {h: i for i, h in enumerate(cleaned)}
    sku_keys = ("sku", "pim sku", "model", "model id", "product id")
    url_keys = ("url", "pdp url", "product url")

    for raw_row in rows:
        values = list(raw_row) + [""] * max(0, len(headers) - len(raw_row))
        row = {headers[i]: values[i] for i in range(len(headers))}

        def get(keys: tuple[str, ...], default: str = "") -> str:
            for key in keys:
                if key in header_map:
                    value = str(values[header_map[key]] or "").strip()
                    if value:
                        return value
            return default

        url = get(url_keys)
        raw_sku = get(sku_keys)
        sku = normalize_sku(raw_sku)
        if not sku and url:
            sku = normalize_sku(url.rstrip("/").split("/")[-1])
        if not sku:
            continue
        products.append(ProductInput(
            request_id=get(("request id", "request_id", "id"), f"R{len(products)+1:03d}"),
            sku=sku,
            category=get(("category", "product category", "제품군")),
            country=get(("country", "market", "국가"), "UK"),
            language=get(("language", "locale", "언어"), "en-GB"),
            url=url,
            product_name=get(("product name", "product_name", "name")),
            key_features=_features_from_row(row),
            source_record_id=get(("source record id", "source_record_id", "record id")),
            original_sku=normalize_original_sku(raw_sku),
        ))
    return products


def parse_tabular_text(text: str) -> list[ProductInput]:
    text = (text or "").strip()
    if not text:
        return []
    flattened = parse_flattened_wide_table(text)
    if flattened:
        return flattened
    url_tokens = [x.strip() for x in re.split(r"[\r\n,;]+", text) if x.strip()]
    if url_tokens and all(_is_url(x) for x in url_tokens):
        products = []
        for url in url_tokens:
            slug = url.rstrip("/").split("/")[-1]
            sku = normalize_sku(slug)
            products.append(ProductInput(
                request_id=f"R{len(products)+1:03d}", sku=sku, category=category_from_url(url),
                country="UK", language="en-GB", url=url,
            ))
        return products
    first_line = text.splitlines()[0]
    delimiter = "\t" if "\t" in first_line else ","
    parsed = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    if not parsed:
        return []

    first = [_clean_header(x) for x in parsed[0]]
    known_headers = {"country", "국가", "#", "ref", "w/m", "ltv", "mnt", "sku", "pim sku", "url", "pdp url"}
    has_header = any(x in known_headers or x.startswith("key_feature") for x in first)
    if has_header:
        return rows_to_products([str(x) for x in parsed[0]], parsed[1:])

    products: list[ProductInput] = []
    for row in parsed:
        cells = [str(x).strip() for x in row if str(x).strip()]
        if not cells:
            continue
        url = next((x for x in cells if _is_url(x)), "")
        sku_value = next((x for x in cells if not _is_url(x)), "")
        sku = normalize_sku(sku_value or (url.rstrip("/").split("/")[-1] if url else ""))
        if sku:
            products.append(ProductInput(
                request_id=f"R{len(products)+1:03d}", sku=sku, url=url,
                category=category_from_url(url), original_sku=normalize_original_sku(sku_value),
            ))
    return products


def parse_flattened_wide_table(text: str) -> list[ProductInput]:
    """Restore a copied wide spreadsheet when the browser clipboard flattens every cell to a line."""
    tokens = []
    for raw in text.splitlines():
        value = raw.strip().lstrip("\ufeff\ufffd")
        if value:
            tokens.append(value)
    normalized = [_clean_header(x) for x in tokens]
    expected = ["country", "#", "ref", "w/m", "ltv", "mnt"]
    header_start = next(
        (i for i in range(max(0, len(tokens) - len(expected) + 1)) if normalized[i:i + len(expected)] == expected),
        None,
    )
    if header_start is None:
        return []

    cursor = header_start + len(expected)
    country = "UK"
    if cursor < len(tokens) and not tokens[cursor].isdigit():
        raw_country = tokens[cursor]
        country = "UK" if raw_country in {"영국", "United Kingdom", "GB"} else raw_country
        cursor += 1

    categories = ["REF", "W/M", "LTV", "MNT"]
    products: list[ProductInput] = []
    while cursor < len(tokens):
        if not tokens[cursor].isdigit():
            cursor += 1
            continue
        cursor += 1  # row number
        if cursor + 4 > len(tokens):
            break
        values = tokens[cursor:cursor + 4]
        if any(value.isdigit() for value in values):
            break
        for category, raw_sku in zip(categories, values):
            sku = normalize_sku(raw_sku)
            if sku:
                products.append(ProductInput(
                    request_id=f"R{len(products)+1:03d}", sku=sku, category=category,
                    country=country, language="en-GB", original_sku=normalize_original_sku(raw_sku),
                ))
        cursor += 4
    return products if len(products) >= 4 else []


def load_products_from_file(filename: str, data: bytes) -> list[ProductInput]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        return rows_to_products([str(x or "") for x in rows[0]], rows[1:])
    return parse_tabular_text(data.decode("utf-8-sig"))


class VisibleTextParser(HTMLParser):
    BLOCKS = {"h1", "h2", "h3", "h4", "li", "p", "dt", "dd", "th", "td", "div"}

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.lines: list[str] = []
        self._buffer: list[str] = []
        self._skip = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in {"script", "style", "svg", "noscript"}:
            self._skip += 1
        if tag in self.BLOCKS:
            self._flush()

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style", "svg", "noscript"} and self._skip:
            self._skip -= 1
        if tag in self.BLOCKS:
            self._flush()

    def handle_data(self, data: str) -> None:
        if not self._skip:
            text = re.sub(r"\s+", " ", data).strip()
            if text:
                self._buffer.append(text)

    def _flush(self) -> None:
        if not self._buffer:
            return
        text = re.sub(r"\s+", " ", " ".join(self._buffer)).strip()
        self._buffer.clear()
        if text and (not self.lines or self.lines[-1] != text):
            self.lines.append(text)

    def close(self) -> None:
        self._flush()
        super().close()


def extract_primary_image(html: str, page_url: str) -> str:
    """Return the PDP primary product image URL.

    Interim policy: prefer the first product-gallery/hero image, then
    Product JSON-LD image, then og:image. Relative URLs are normalized.
    Logos, icons, badges, video posters and tiny placeholders are excluded.
    """
    decoded = html.replace("\\/", "/")

    def normalize(value: str) -> str:
        value = re.sub(r"&amp;", "&", value or "").strip().strip('"\'')
        if not value or value.startswith(("data:", "blob:")):
            return ""
        return urljoin(page_url, value)

    def allowed(url: str) -> bool:
        low = url.casefold()
        if not low.startswith(("http://", "https://")):
            return False
        blocked = (
            "logo", "icon", "badge", "award", "rating", "review", "paypal",
            "klarna", "placeholder", "spinner", "loading", "sprite", "favicon",
            "product-information", "energy-label", "video-poster", "youtube",
        )
        if any(term in low for term in blocked):
            return False
        if low.endswith((".svg", ".gif")):
            return False
        return bool(re.search(r"\.(?:jpe?g|png|webp)(?:[?#].*)?$", low) or "image" in low)

    def largest_srcset(value: str) -> str:
        candidates=[]
        for part in value.split(','):
            bits=part.strip().split()
            if not bits:
                continue
            score=0
            if len(bits)>1:
                m=re.match(r"([0-9.]+)(w|x)$", bits[-1], re.I)
                if m:
                    score=float(m.group(1))*(10000 if m.group(2).lower()=='x' else 1)
            candidates.append((score,bits[0]))
        return max(candidates, default=(0,""))[1]

    # 1) Product gallery / hero assets in source order. Prefer srcset maximum.
    gallery_patterns = (
        r'<(?:img|source)[^>]+(?:class|data-testid|id)=["\'][^"\']*(?:gallery|hero|product-image|pdp-image)[^"\']*["\'][^>]+(?:srcset|data-srcset)=["\']([^"\']+)["\']',
        r'<(?:img|source)[^>]+(?:class|data-testid|id)=["\'][^"\']*(?:gallery|hero|product-image|pdp-image)[^"\']*["\'][^>]+(?:src|data-src|data-original)=["\']([^"\']+)["\']',
        r'<(?:img|source)[^>]+(?:srcset|data-srcset)=["\']([^"\']+)["\'][^>]+(?:class|data-testid|id)=["\'][^"\']*(?:gallery|hero|product-image|pdp-image)[^"\']*["\']',
        r'<(?:img|source)[^>]+(?:src|data-src|data-original)=["\']([^"\']+)["\'][^>]+(?:class|data-testid|id)=["\'][^"\']*(?:gallery|hero|product-image|pdp-image)[^"\']*["\']',
    )
    for idx, pattern in enumerate(gallery_patterns):
        for value in re.findall(pattern, decoded, flags=re.I | re.S):
            candidate = largest_srcset(value) if idx in {0,2} else value
            url=normalize(candidate)
            if allowed(url):
                return url

    # 2) Product JSON-LD image. Accept string or first array entry.
    for block in re.findall(r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>', decoded, flags=re.I | re.S):
        if not re.search(r'"@type"\s*:\s*(?:"Product"|\[[^\]]*"Product")', block, flags=re.I | re.S):
            continue
        match=re.search(r'"image"\s*:\s*(?:\[\s*)?"([^"\]]+)"', block, flags=re.I | re.S)
        if match:
            url=normalize(match.group(1))
            if allowed(url):
                return url

    # 3) Social metadata fallback.
    meta_patterns=(
        r'<meta[^>]+property=["\']og:image(?::secure_url)?["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image(?::secure_url)?["\']',
        r'<meta[^>]+name=["\']twitter:image["\'][^>]+content=["\']([^"\']+)["\']',
    )
    for pattern in meta_patterns:
        match=re.search(pattern, decoded, flags=re.I | re.S)
        if match:
            url=normalize(match.group(1))
            if allowed(url):
                return url
    return ""


def extract_msrp(html: str) -> str:
    """Extract a numeric MSRP/list price from PDP HTML.

    Interim policy: prefer explicit MSRP/RRP/list-price fields. For UK PDPs
    where the visible price block contains sale, list and member prices,
    use the highest visible GBP product price as the MSRP. Currency is not
    included in the returned feed value.
    """
    decoded = html.replace("\\/", "/")

    def _amount(value: str) -> float | None:
        try:
            amount = float(value.replace(",", ""))
        except (TypeError, ValueError):
            return None
        return amount if 0 < amount < 100000 else None

    # 1) Explicit MSRP-type fields are the most trusted source.
    explicit_patterns = (
        r'"(?:msrpPrice|msrp_price|msrp|rrp|listPrice)"\s*:\s*"?([0-9][0-9,]*(?:\.[0-9]{1,2})?)"?',
        r'(?:data-msrp|data-rrp|data-list-price)=["\']([0-9][0-9,]*(?:\.[0-9]{1,2})?)["\']',
    )
    explicit: list[float] = []
    for pattern in explicit_patterns:
        for value in re.findall(pattern, decoded, flags=re.IGNORECASE):
            amount = _amount(value)
            if amount is not None:
                explicit.append(amount)
    if explicit:
        return f"{max(explicit):.2f}"

    # 2) UK PDP fallback: visible commerce blocks often contain sale price,
    # list price and member price together. The highest GBP product price is
    # used as the interim MSRP; Save/discount amounts are excluded.
    parser = VisibleTextParser()
    parser.feed(decoded)
    parser.close()
    visible = " ".join(parser.lines)
    visible = re.sub(r"Save\s*£\s*[0-9][0-9,]*(?:\.[0-9]{1,2})?", " ", visible, flags=re.I)
    visible = re.sub(r"(?:Trade-?Up|cashback|voucher)[^£]{0,30}£\s*[0-9][0-9,]*(?:\.[0-9]{1,2})?", " ", visible, flags=re.I)
    gbp_values: list[float] = []
    for value in re.findall(r"£\s*([0-9][0-9,]*(?:\.[0-9]{1,2})?)", visible):
        amount = _amount(value)
        if amount is not None and amount >= 50:
            gbp_values.append(amount)
    if gbp_values:
        return f"{max(gbp_values):.2f}"

    # 3) Standard structured selling-price fallbacks.
    trusted_patterns = (
        r'"(?:regularPrice|sellingPrice|salePrice|finalPrice|offerPrice)"\s*:\s*"?([0-9][0-9,]*(?:\.[0-9]{1,2})?)"?',
        r'<meta[^>]+(?:itemprop=["\']price["\']|property=["\']product:price:amount["\'])[^>]+content=["\']([0-9][0-9,]*(?:\.[0-9]{1,2})?)["\']',
        r'<meta[^>]+content=["\']([0-9][0-9,]*(?:\.[0-9]{1,2})?)["\'][^>]+(?:itemprop=["\']price["\']|property=["\']product:price:amount["\'])',
    )
    for pattern in trusted_patterns:
        match = re.search(pattern, decoded, flags=re.IGNORECASE)
        if match:
            amount = _amount(match.group(1))
            if amount is not None:
                return f"{amount:.2f}"

    # JSON-LD Offer price is accepted only when accompanied by a currency.
    for block in re.findall(r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>', decoded, flags=re.I | re.S):
        if not re.search(r'"priceCurrency"\s*:\s*"[A-Z]{3}"', block, flags=re.I):
            continue
        match = re.search(r'"price"\s*:\s*"?([0-9][0-9,]*(?:\.[0-9]{1,2})?)"?', block, flags=re.I)
        if match:
            amount = _amount(match.group(1))
            if amount is not None:
                return f"{amount:.2f}"
    return ""

def taxonomy_config_path() -> Path:
    """Return the live Excel config path; the environment variable supports shared locations."""
    override = os.getenv("VF_TAXONOMY_CONFIG", "").strip()
    return Path(override) if override else Path(__file__).with_name("feed_taxonomy_config.xlsx")


def _sheet_rows(workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"설정 Excel에 {sheet_name} 시트가 없습니다.")
    sheet = workbook[sheet_name]
    values = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(values)]
    return [dict(zip(headers, row)) for row in values if any(value not in {None, ""} for value in row)]


def load_taxonomy_config() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Read the saved workbook on every generation so changes apply without server restart."""
    from openpyxl import load_workbook
    path = taxonomy_config_path()
    if not path.exists():
        raise FileNotFoundError(f"Taxonomy 설정 파일을 찾을 수 없습니다: {path}")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except PermissionError as exc:
        raise RuntimeError("Taxonomy Excel을 저장하고 닫은 뒤 다시 실행해 주세요.") from exc
    try:
        return (
            _sheet_rows(workbook, "Market_Category"),
            _sheet_rows(workbook, "Intent_Label"),
            _sheet_rows(workbook, "Feature_Intent_Rule"),
            _sheet_rows(workbook, "Feed_Exclusion_Rule"),
            _sheet_rows(workbook, "Copy_Template"),
            _sheet_rows(workbook, "Product_Subtype"),
        )
    finally:
        workbook.close()


def _active(row: dict[str, Any]) -> bool:
    enabled = str(row.get("enabled", "Y") or "Y").strip().upper() in {"Y", "YES", "TRUE", "1"}
    status = str(row.get("validation_status", "") or "").strip().lower()
    return enabled and status != "rejected"


def _priority(row: dict[str, Any]) -> int:
    try:
        return int(row.get("priority", 100) or 100)
    except (TypeError, ValueError):
        return 100


def _select_mapping(rows: list[dict[str, Any]], market: str, category: str, intent: str | None = None) -> dict[str, Any] | None:
    market = market.strip().upper()
    category = category.strip().upper()
    ranked: list[tuple[int, int, dict[str, Any]]] = []
    for row in rows:
        if not _active(row):
            continue
        row_market = str(row.get("market", "*") or "*").strip().upper()
        row_category = str(row.get("category_code", "*") or "*").strip().upper()
        if row_market not in {market, "*"} or row_category not in {category, "*"}:
            continue
        if intent is not None and str(row.get("internal_intent", "") or "").strip() != intent:
            continue
        specificity = (2 if row_market == market else 0) + (1 if row_category == category else 0)
        ranked.append((-specificity, _priority(row), row))
    return min(ranked, default=None, key=lambda item: (item[0], item[1]))[2] if ranked else None


def product_subtype_label(product: ProductInput, rows: list[dict[str, Any]]) -> tuple[str, str] | None:
    market = product.country.strip().upper()
    category = product.category.strip().upper()
    haystacks = (product.url.casefold(), product.product_name.casefold())
    ranked = []
    for row in rows:
        if not _active(row):
            continue
        row_market = str(row.get("market", "*") or "*").upper()
        row_category = str(row.get("category_code", "*") or "*").upper()
        if row_market not in {market, "*"} or row_category not in {category, "*"}:
            continue
        url_terms = _terms(row.get("url_pattern"))
        name_terms = _terms(row.get("product_name_pattern"))
        url_match = any(term in haystacks[0] for term in url_terms) if url_terms else False
        name_match = any(term in haystacks[1] for term in name_terms) if name_terms else False
        if not (url_match or name_match):
            continue
        specificity = (2 if row_market == market else 0) + (1 if row_category == category else 0)
        ranked.append((-specificity, _priority(row), row))
    if not ranked:
        return None
    row = min(ranked, key=lambda item: (item[0], item[1]))[2]
    label = str(row.get("feed_category_label", "") or "").strip()
    short = str(row.get("short_label", label) or label).strip()
    return (label, short) if label else None


def product_title(product: ProductInput, evidence: Evidence, limit: int) -> str:
    """Build a title from the live market/category and intent Excel mappings."""
    category_rows, intent_rows, _, _, _, subtype_rows = load_taxonomy_config()
    category_row = _select_mapping(category_rows, product.country, product.category)
    intent_row = _select_mapping(intent_rows, product.country, product.category, evidence.intent_candidate)
    if not category_row:
        raise ValueError(f"제품군 Mapping이 없습니다: {product.country}/{product.category}")
    category_label = str(category_row.get("feed_category_label", "") or "").strip()
    short_category = str(category_row.get("short_label", category_label) or category_label).strip()
    subtype = product_subtype_label(product, subtype_rows)
    if subtype:
        category_label, short_category = subtype
    # A newly classified intent must not cancel every feed for the product.
    # Use the internal intent as a review-safe fallback until the Excel mapping is added.
    if intent_row:
        intent_label = str(intent_row.get("title_intent_label", "") or "").strip()
        short_intent = str(intent_row.get("short_intent_label", intent_label) or intent_label).strip()
    else:
        fallback_labels = {
            "Product Benefit": ("Key Benefit", "Benefit"),
            "Capacity": ("More Space", "Space"),
            "Design": ("Design", "Design"),
        }
        intent_label, short_intent = fallback_labels.get(
            evidence.intent_candidate,
            (evidence.intent_candidate, evidence.intent_candidate),
        )
    candidates = (
        f"LG {category_label}: {intent_label}",
        f"LG {short_category}: {intent_label}",
        f"LG {short_category}: {short_intent}",
    )
    for title in candidates:
        if len(title) <= limit:
            return title
    return fit_text(candidates[-1], limit)


class PDPExtractor:
    STOP_MARKERS = {
        "more", "deals & offers", "gallery", "gallery ar", "features specs reviews support",
        "zoom in zoom out", "previous slide", "next slide", "what's to love about lg fridge freezers?",
    }

    def fetch(self, url: str, timeout: int = 25) -> str:
        request = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (compatible; LG-Virtual-Feed-Discovery/0.1)",
                "Accept-Language": "en-GB,en;q=0.9",
            },
        )
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                charset = response.headers.get_content_charset() or "utf-8"
                return response.read().decode(charset, errors="replace")
        except urllib.error.HTTPError as exc:
            if exc.code not in {401, 403, 429}:
                raise
            return dump_dom(url, timeout=max(timeout, 40))

    def extract(self, product: ProductInput, html: str) -> tuple[str, list[Evidence], list[Issue]]:
        sales_model, market_sku = extract_product_codes(html)
        if sales_model:
            product.sku = sales_model
        if market_sku:
            product.original_sku = market_sku

        parser = VisibleTextParser()
        parser.feed(html)
        parser.close()
        lines = parser.lines
        issues: list[Issue] = []
        product.price = extract_msrp(html)
        if not product.image_link:
            product.image_link = extract_primary_image(html, product.url)

        current_key = re.sub(r"[^A-Z0-9]", "", product.sku.upper())
        model_candidates: list[str] = []
        for line in lines:
            if not re.fullmatch(r"[A-Za-z0-9-]{6,24}", line.strip()):
                continue
            candidate = normalize_sku(line)
            candidate_key = re.sub(r"[^A-Z0-9]", "", candidate)
            if candidate_key and (
                current_key.startswith(candidate_key) or candidate_key.startswith(current_key)
            ) and abs(len(current_key) - len(candidate_key)) <= 5:
                model_candidates.append(candidate)
        if model_candidates:
            product.sku = min(model_candidates, key=len)

        title = product.product_name
        if not title:
            for line in lines:
                if product.sku.replace("-", "") in line.upper().replace("-", "") and 8 < len(line) < 180:
                    title = line
                    break

        features: list[str] = []
        feature_headings = {"key features", "principais recursos", "principais características", "principais caracteristicas"}
        starts = [i for i, line in enumerate(lines) if line.strip().lower() in feature_headings]
        for start in starts:
            candidate: list[str] = []
            for line in lines[start + 1:start + 14]:
                clean = line.strip()
                if clean.lower() in self.STOP_MARKERS:
                    break
                if len(clean) < 8 or clean.lower() in {"yes", "no"}:
                    continue
                if clean.startswith(("Image:", "Front view", "LG Product Service")):
                    continue
                candidate.append(clean)
                if len(candidate) >= 7:
                    break
            if len(candidate) > len(features):
                features = candidate

        if not features:
            issues.append(Issue(product.sku, "extract", "error", "NO_KEY_FEATURES", "Key Features를 찾지 못했습니다."))

        evidence: list[Evidence] = []
        for idx, feature in enumerate(dict.fromkeys(features), 1):
            reason = exclusion_reason(feature, product.category, product.country)
            if reason:
                issues.append(Issue(product.sku, "qc", "info", reason, f"Feed 제외: {feature}"))
                continue
            evidence.append(Evidence(
                claim_id=f"{product.sku}-C{idx:02d}",
                sku=product.sku,
                category=product.category,
                source_type="PDP",
                source_record_id=product.url,
                source_section="Key Features",
                source_text=feature,
                intent_candidate=classify_intent(feature, product.category, product.country),
                validation_status="Needs Review",
                notes="PDP Key Feature에서 자동 추출. Key Specs/각주 교차 확인 전.",
            ))
        return title, evidence, issues


class LGSitemapResolver:
    EXCLUDED_PATHS = ("/support/", "/tncs/", "/promotion/", "/microsite", "/business/")
    CATEGORY_HINTS = {
        "REF": ("/fridge-freezers/", "/geladeiras/", "/refrigeradores/"),
        "W/M": ("/laundry/", "/lavanderia/", "/maquinas-de-lavar/"),
        "LTV": ("/tvs-soundbars/", "/televisions/", "/tv/", "/tvs/"),
        "MNT": ("/monitors/", "/monitores/"),
    }
    _urls: dict[str, list[str]] = {}
    _lock = threading.Lock()

    @classmethod
    def _load_urls(cls, country: str) -> list[str]:
        config = country_config(country)
        if not config or not config.get("enabled"):
            raise ValueError(f"지원하지 않는 국가입니다: {country}")
        site_code = config["lg_site_code"].lower()
        if site_code in cls._urls:
            return cls._urls[site_code]
        with cls._lock:
            if site_code in cls._urls:
                return cls._urls[site_code]
            request = urllib.request.Request(
                config["sitemap_url"],
                headers={"User-Agent": "Mozilla/5.0 (compatible; LG-Virtual-Feed-Discovery/0.1)"},
            )
            with urllib.request.urlopen(request, timeout=25) as response:
                xml = response.read().decode("utf-8", errors="replace")
            pattern = rf"<loc>(https://www\.lg\.com/{re.escape(site_code)}/[^<]+)</loc>"
            cls._urls[site_code] = re.findall(pattern, xml, flags=re.IGNORECASE)
            return cls._urls[site_code]

    @classmethod
    def select_url(cls, sku: str, category: str, urls: Iterable[str]) -> str:
        sku_key = re.sub(r"[^a-z0-9]", "", normalize_sku(sku).lower())
        hints = cls.CATEGORY_HINTS.get(category.upper(), ())
        candidates: list[tuple[int, str]] = []
        for url in urls:
            lowered = url.lower().split("?", 1)[0]
            if any(blocked in lowered for blocked in cls.EXCLUDED_PATHS):
                continue
            slug = lowered.rstrip("/").rsplit("/", 1)[-1]
            slug_key = re.sub(r"[^a-z0-9]", "", slug)
            if not slug_key.startswith(sku_key):
                continue
            suffix = slug_key[len(sku_key):]
            if len(suffix) > 8:
                continue
            score = 100 if not suffix else 94 if suffix == "1" else 75
            if any(hint in lowered for hint in hints):
                score += 20
            if "bundle" in lowered or suffix.startswith(("ms", "fdv")):
                score -= 15
            candidates.append((score, url.split("?", 1)[0]))
        return max(candidates, default=(0, ""), key=lambda item: (item[0], -len(item[1])))[1]

    @classmethod
    def resolve(cls, country: str, sku: str, category: str) -> str:
        candidates: list[str] = []
        try:
            candidates = cls._load_urls(country)
        except urllib.error.HTTPError as exc:
            if exc.code not in {401, 403, 429}:
                raise
        selected = cls.select_url(sku, category, candidates)
        # Some country sitemaps return an index or omit PDP URLs without an HTTP error.
        # In that case, always fall back to official-domain browser search.
        if not selected:
            config = country_config(country) or {}
            search_candidates = search_lg_pdp(str(config.get("lg_site_code", "")).lower(), sku)
            selected = cls.select_url(sku, category, search_candidates)
        return selected


INTENT_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("Easy Access", ("instaview", "door-in-door", "easy access", "quick access")),
    ("Freshness", ("fresh", "linearcooling", "doorcooling", "naturefresh", "temperature")),
    ("Hygiene Care", ("uvnano", "bacteria", "allergen", "hygiene", "steam")),
    ("Smart Convenience", ("thinq", "smart learner", "wi-fi", "wifi", "app", "smartphone", "ai ")),
    ("Speed & Time Saving", ("turbowash", "minutes", "quick", "speed", "refresh rate", "response time")),
    ("Energy Saving", ("save energy", "energy efficient", "energy saving", "consumption", "eficiência energética", "eficiencia energetica", "economia de energia", "econômico", "economico")),
    ("Reliability", ("warranty", "compressor warranty", "reliable", "garantia", "durável", "duravel")),
    ("Quiet & Reliable", ("quiet", "low noise", "silencioso")),
    ("Capacity", ("capacity", "storage", " kg", " litre", "liter", "635l", "capacidade", " litros", "451l", "462l")),
    ("Picture Quality", ("oled", "brightness", "hdr", "resolution", "dci-p3", "contrast", "colour fidelity", "black level")),
    ("Gaming Performance", ("gaming", "hz", "gtg", "g-sync", "freesync", "dual-mode")),
    ("Connectivity", ("usb", "hdmi", "displayport", "connect")),
    ("Design", ("design", "sleek", "finish", "style", "bezel")),
]


def _terms(value: Any) -> list[str]:
    return [term.strip().lower() for term in str(value or "").split("|") if term.strip()]


def classify_intent(text: str, category: str = "", market: str = "*") -> str:
    lowered = f" {text.lower()} "
    # Deterministic UK monitor precedence for specs that overlap across intents.
    if market.upper() == "UK" and category.upper() == "MNT":
        if any(term in lowered for term in (" dqhd", " dual qhd", " 5120x1440")):
            return "Resolution"
        if any(term in lowered for term in (" contrast ratio", " 2,000:1", " 2000:1", " 1.5m:1")):
            return "Contrast"
        if any(term in lowered for term in (" webos", " airplay 2", " screen share", " magic remote")):
            return "Smart Convenience"
        if any(term in lowered for term in (" stand with wheels", " portrait mode", " swivel", " adjustable stand")):
            return "Design"
        if any(term in lowered for term in (" 165hz", " 330hz", " 240hz", " 144hz", " refresh rate", " 0.03ms", " 1ms mbr", " g-sync", " freesync", " dual-mode")):
            return "Gaming Performance"
        if any(term in lowered for term in (" dci-p3", " srgb", " adobe rgb", " colour calibrated", " colour accuracy", " real 10-bit")):
            return "Colour Accuracy"
        if any(term in lowered for term in (" displayhdr", " true black", " hdr 400", " hdr600", " hdr 600")):
            return "HDR Experience"
        if any(term in lowered for term in (" thunderbolt", " usb-c", " usb type-c", " displayport", " dp 2.1", " hdmi", " pd90w", " pd 90w", " pd 65w")):
            return "Connectivity"
        if any(term in lowered for term in (" ai sound", " ai scene optimisation", " ai scene optimization", " ai upscaling")):
            return "AI Experience"
        if any(term in lowered for term in (" 1500nits", " 1300 nits", " peak brightness", " tandem oled")):
            return "Brightness"
        if any(term in lowered for term in (" 5k2k", " 6k", " 4k", " wqhd", " qhd", " 5120x2160", " 6144 x 3456", " 6144x3456", " 3440x1440", " 2560 x 1440", " 143 ppi", " 224ppi")):
            return "Resolution"

    # Deterministic UK TV precedence for overlapping OLED and AI claims.
    if market.upper() == "UK" and category.upper() == "LTV":
        if any(term in lowered for term in (" 5 year limited panel warranty", " 5-year limited panel warranty")):
            return "Reliability"
        if any(term in lowered for term in (" 165hz", " 144hz", " 120hz", " g-sync", " freesync", " gameplay")):
            return "Gaming Performance"
        if any(term in lowered for term in (" x3.9", " 3.9 higher peak brightness", " 3x brighter", " brightness booster", " peak brightness")):
            return "Brightness"
        if any(term in lowered for term in (" perfect black", " true black", " deeper contrast", " stunning contrast")):
            return "Contrast"
        if any(term in lowered for term in (" webos", " google gemini", " microsoft copilot", " ai magic remote", " ai button", " voice controls")):
            return "Smart Convenience"
        if any(term in lowered for term in (" hyper radiant colour", " 100% colour fidelity", " 100% colour volume", " lifelike colours", " perfect colour")):
            return "Picture Quality"
        if any(term in lowered for term in (" alpha 11 ai processor", " ai upscaled", " ai picture", " upscale")):
            return "AI Experience"

    # Deterministic UK laundry precedence for overlapping AI, care and warranty claims.
    if market.upper() == "UK" and category.upper() == "W/M":
        if any(term in lowered for term in (" ai direct drive", " ai dd", " aidd", " ai wash", " fabric care", " fabric characteristics", " best motion", " smart pairing")):
            return "Fabric Care"
        if any(term in lowered for term in (" thinq", " wi-fi", " wifi", " remote control", " monitoring", " download of new cycles")):
            return "Smart Convenience"
        if any(term in lowered for term in (" auto cleaning condenser", " self-cleans", " condenser each cycle")):
            return "Easy Maintenance"
        if any(term in lowered for term in (" dual inverter heat pump", " inverter compressor", " inverter motor", " 10 year warranty", " 10-year warranty")):
            return "Reliability"
        if any(term in lowered for term in (" allergy care", " allergens", " dust mites", " steam")):
            return "Hygiene Care"

    # Deterministic UK refrigerator precedence for overlapping feature text.
    if market.upper() == "UK" and category.upper() == "REF":
        if any(term in lowered for term in (" smart inverter compressor", " inverter compressor", " 10-year", " 10 year", " warranty")):
            return "Reliability"
        if any(term in lowered for term in (" plumbed fridge", " water line", " filtered water", " internal filter")):
            return "Water & Ice"
        if any(term in lowered for term in (" large capacity", " maximum storage", " max storage", " storage designed", " bags of shopping")):
            return "Capacity"
        if any(term in lowered for term in (" thinq", " smart learner", " ai saving mode")):
            return "Smart Convenience"
        if any(term in lowered for term in (" instaview", " door-in-door")):
            return "Easy Access"
        if any(term in lowered for term in (" doorcooling", " linear cooling", " linearcooling", " naturefresh", " freshbalancer", " freshconverter")):
            return "Freshness"

    # Deterministic BR TV precedence for overlapping picture claims.
    # This prevents "preto perfeito" / explicit contrast evidence from being
    # absorbed by the broader Picture Quality rule.
    if market.upper() == "BR" and category.upper() == "LTV":
        if any(term in lowered for term in (
            " precision dimming", " mini led", " preto perfeito",
            " contraste mais profundo", " contraste excepcional",
        )):
            return "Contrast"
    try:
        _, _, rules, _, _, _ = load_taxonomy_config()
        ranked = []
        for row in rules:
            if not _active(row):
                continue
            row_market = str(row.get("market", "*") or "*").upper()
            row_category = str(row.get("category_code", "*") or "*").upper()
            if row_market not in {market.upper(), "*"} or row_category not in {category.upper(), "*"}:
                continue
            if any(term in lowered for term in _terms(row.get("negative_pattern"))):
                continue
            if any(term in lowered for term in _terms(row.get("keyword_pattern"))):
                specificity = (2 if row_market == market.upper() else 0) + (1 if row_category == category.upper() else 0)
                ranked.append((-specificity, _priority(row), str(row.get("internal_intent", "Product Benefit"))))
        if ranked:
            return min(ranked, key=lambda item: (item[0], item[1]))[2]
    except (FileNotFoundError, ValueError):
        pass
    for intent, terms in INTENT_RULES:
        if any(term in lowered for term in terms):
            return intent
    return "Product Benefit"


def exclusion_reason(text: str, category: str = "", market: str = "*") -> str:
    lowered = text.lower()
    try:
        _, _, _, rules, _, _ = load_taxonomy_config()
        ranked = []
        for row in rules:
            if not _active(row):
                continue
            row_market = str(row.get("market", "*") or "*").upper()
            row_category = str(row.get("category_code", "*") or "*").upper()
            if row_market not in {market.upper(), "*"} or row_category not in {category.upper(), "*"}:
                continue
            if any(term in lowered for term in _terms(row.get("keyword_pattern"))):
                specificity = (2 if row_market == market.upper() else 0) + (1 if row_category == category.upper() else 0)
                ranked.append((-specificity, _priority(row), str(row.get("reason_code", "EXCLUDED"))))
        if ranked:
            return min(ranked, key=lambda item: (item[0], item[1]))[2]
    except (FileNotFoundError, ValueError):
        pass
    return ""


TITLE_BY_INTENT = {
    "Easy Access": "Everyday Access Made Easy",
    "Freshness": "Freshness That Lasts Longer",
    "Hygiene Care": "Care for Everyday Hygiene",
    "Smart Convenience": "Smarter Control, Every Day",
    "Speed & Time Saving": "More Done in Less Time",
    "Energy Saving": "Designed to Save Energy",
    "Quiet & Reliable": "Quiet, Efficient Performance",
    "Capacity": "Room for More",
    "Picture Quality": "See Every Detail",
    "Gaming Performance": "Made for Faster Gaming",
    "Connectivity": "Connect with Ease",
    "Design": "Designed to Fit Your Style",
    "Product Benefit": "Made for Everyday Life",
}


def fit_text(text: str, limit: int) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= limit:
        return text
    shortened = text[: limit + 1].rsplit(" ", 1)[0].rstrip(" ,;:-")
    return shortened if shortened else text[:limit]


_DANGLING_END = re.compile(r"\b(?:and|or|with|in|on|for|to|from|of|the|a|an|up to|can|may|including)$", re.I)


def is_complete_copy(text: str) -> bool:
    clean = re.sub(r"\s+", " ", text).strip().rstrip(".,!?;:")
    return bool(clean) and not _DANGLING_END.search(clean)


def localize_evidence_copy(source_text: str, market: str) -> str:
    """Localize common PDP feature labels without inventing product claims."""
    text = re.sub(r"\s+", " ", source_text).strip()
    if market.upper() != "BR":
        return text
    exact = {
        "smart inverter compressor": "Compressor Smart Inverter",
        "gaveta fresh zone": "Gaveta Fresh Zone",
        "design premium": "Design premium",
        "door cooling+": "Resfriamento DoorCooling+",
        "door cooling⁺": "Resfriamento DoorCooling+",
        "multi air flow": "Fluxo de ar multidirecional",
        "alta eficiência energética a": "Alta eficiência energética classe A",
    }
    lowered = text.casefold()
    if lowered in exact:
        return exact[lowered]
    replacements = (
        (r"\benergy saving\b", "economia de energia"),
        (r"\bhigh energy efficiency\b", "alta eficiência energética"),
        (r"\breliability\b", "confiabilidade"),
        (r"\bfreshness\b", "frescor"),
        (r"\bpremium design\b", "design premium"),
        (r"\blarge capacity\b", "alta capacidade"),
    )
    for pattern, replacement in replacements:
        text = re.sub(pattern, replacement, text, flags=re.I)
    return text


def copy_template_for(text: str, market: str, category: str, intent: str) -> str:
    try:
        _, _, _, _, rows, _ = load_taxonomy_config()
    except (FileNotFoundError, ValueError):
        return ""
    low=text.casefold(); matches=[]
    for row in rows:
        if not _active(row): continue
        rm=str(row.get("market","*") or "*").upper(); rc=str(row.get("category_code","*") or "*").upper()
        if rm not in {market.upper(),"*"} or rc not in {category.upper(),"*"} or str(row.get("internal_intent","") or "") != intent: continue
        if not any(term in low for term in _terms(row.get("keyword_pattern"))): continue
        required = _terms(row.get("required_pattern"))
        if required and not all(term in low for term in required): continue
        spec=(2 if rm==market.upper() else 0)+(1 if rc==category.upper() else 0)
        matches.append((-spec,_priority(row),str(row.get("description_template","") or "").strip()))
    return min(matches,default=(0,0,""),key=lambda z:(z[0],z[1]))[2]


def rule_body(source_text: str, limit: int, market: str = "", category: str = "", intent: str = "") -> str:
    """Return a complete evidence-grounded clause; never hard-cut mid-phrase."""
    text = localize_evidence_copy(source_text, market)
    governed = copy_template_for(source_text, market, category, intent)
    if governed and len(governed) <= limit and is_complete_copy(governed): return governed
    lowered = text.lower()
    templates = []
    if "uvnano" in lowered and "bacteria" in lowered:
        templates.append("UVnano™ dispenser helps eliminate bacteria.")
    if "instaview" in lowered and "door-in-door" in lowered:
        templates.append("See inside easily with InstaView™ Door-in-Door™.")
    if "naturefresh" in lowered and "fresh" in lowered:
        templates.append("NatureFRESH™ helps keep food fresher for longer.")
    if ("detergent" in lowered or "softener" in lowered) and ("precise" in lowered or "dose" in lowered or "amount" in lowered):
        templates.append("Dispenses the right amount of detergent for each load.")
    capacity = re.search(r"\b(\d+(?:\.\d+)?)\s*kg\b", lowered)
    if capacity and "capacity" in lowered:
        templates.append(f"Handles up to {capacity.group(1)}kg loads with large capacity.")
    warranty = re.search(r"\b(\d+)\s*[- ]?year\b", lowered)
    if warranty and "warranty" in lowered:
        subject = "panel" if "panel" in lowered else "compressor" if "compressor" in lowered else "product"
        templates.append(f"Includes a {warranty.group(1)}-year limited {subject} warranty.")
    for candidate in templates:
        if len(candidate) <= limit and is_complete_copy(candidate):
            return candidate
    weak_label = len(text.split()) <= 4 or text.rstrip().endswith((':', '™', '®', '+™'))
    if len(text) <= limit and is_complete_copy(text) and not weak_label:
        return text
    # Prefer a complete punctuation-delimited clause from the evidence.
    for clause in re.split(r"(?<=[.!?;])\s+|\s+[–—]\s+", text):
        clause = clause.strip(" ,;:-")
        if clause and len(clause) <= limit and is_complete_copy(clause):
            return clause + ("." if clause[-1] not in ".!?" else "")
    # Conservative reductions: remove trailing elaboration only at semantic boundaries.
    for separator in (", ", " with ", " featuring ", " including ", " and "):
        head = text.split(separator, 1)[0].strip(" ,;:-")
        if head and len(head) <= limit and is_complete_copy(head):
            return head + ("." if head[-1] not in ".!?" else "")
    return ""


class RuleBasedGenerator:
    name = "rules"

    def generate(self, product: ProductInput, evidence: list[Evidence], max_feeds: int, title_limit: int, body_limit: int) -> list[Feed]:
        selected: list[Evidence] = []
        seen: set[str] = set()
        for item in evidence:
            if item.intent_candidate in seen:
                continue
            selected.append(item)
            seen.add(item.intent_candidate)
            if len(selected) >= max_feeds:
                break

        feeds: list[Feed] = []
        for item in selected:
            if item.intent_candidate == "Product Benefit":
                continue
            title = product_title(product, item, title_limit)
            body = rule_body(item.source_text, body_limit, product.country, product.category, item.intent_candidate)
            if not body or not is_complete_copy(body):
                continue
            idx = len(feeds) + 1
            feeds.append(Feed(
                sku=product.sku, virtual_sku=f"{product.original_sku or product.sku}_{idx}",
                category=product.category, country=product.country, language=product.language,
                feed_no=idx, intent=item.intent_candidate, brand_title=title, title_chars=len(title),
                brand_body_copy=body, body_chars=len(body), evidence_ids=item.claim_id,
                validation_status="Needs Review",
                review_notes="규칙 기반 초안입니다. 원문을 축약했으므로 사람 검토가 필요합니다.",
            ))
        return feeds


class OpenAIGenerator:
    name = "openai"

    def __init__(self, api_key: str, model: str) -> None:
        self.api_key = api_key
        self.model = model

    def generate(self, product: ProductInput, evidence: list[Evidence], max_feeds: int, title_limit: int, body_limit: int) -> list[Feed]:
        schema = {
            "type": "object",
            "properties": {
                "feeds": {
                    "type": "array", "maxItems": max_feeds,
                    "items": {
                        "type": "object",
                        "properties": {
                            "intent": {"type": "string"},
                            "brand_title": {"type": "string"},
                            "brand_body_copy": {"type": "string"},
                            "evidence_ids": {"type": "array", "items": {"type": "string"}},
                            "review_notes": {"type": "string"},
                        },
                        "required": ["intent", "brand_title", "brand_body_copy", "evidence_ids", "review_notes"],
                        "additionalProperties": False,
                    },
                }
            },
            "required": ["feeds"], "additionalProperties": False,
        }
        evidence_payload = [{"claim_id": e.claim_id, "intent_candidate": e.intent_candidate, "source_text": e.source_text} for e in evidence]
        instructions = (
            "Create evidence-grounded UK English virtual advertising feeds for one product. "
            "Use only explicit claims in the evidence. Do not invent benefits, numbers, comparisons, certifications, promotions, or applicability. "
            "Each feed must express a meaningfully different purchase intent. Return fewer feeds when evidence is insufficient. "
            f"Count spaces and punctuation: title <= {title_limit} characters; body <= {body_limit} characters. "
            "Every feed must cite one or more claim_id values. Flag any caveat in review_notes."
        )
        payload = {
            "model": self.model,
            "input": [
                {"role": "system", "content": [{"type": "input_text", "text": instructions}]},
                {"role": "user", "content": [{"type": "input_text", "text": json.dumps({"sku": product.sku, "category": product.category, "country": product.country, "language": product.language, "evidence": evidence_payload}, ensure_ascii=False)}]},
            ],
            "reasoning": {"effort": "low"},
            "text": {"format": {"type": "json_schema", "name": "virtual_feeds", "strict": True, "schema": schema}},
            "max_output_tokens": 2500,
        }
        request = urllib.request.Request(
            "https://api.openai.com/v1/responses",
            data=json.dumps(payload).encode("utf-8"),
            headers={"Authorization": f"Bearer {self.api_key}", "Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=90) as response:
                raw = json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"OpenAI API {exc.code}: {detail[:500]}") from exc

        output_text = ""
        for item in raw.get("output", []):
            if item.get("type") == "message":
                for content in item.get("content", []):
                    if content.get("type") == "output_text":
                        output_text += content.get("text", "")
        parsed = json.loads(output_text)
        valid_ids = {e.claim_id for e in evidence}
        feeds: list[Feed] = []
        for idx, item in enumerate(parsed.get("feeds", [])[:max_feeds], 1):
            title = item["brand_title"].strip()
            body = item["brand_body_copy"].strip()
            ids = [x for x in item["evidence_ids"] if x in valid_ids]
            status = "Supported" if ids and len(title) <= title_limit and len(body) <= body_limit else "Needs Review"
            feeds.append(Feed(
                sku=product.sku, virtual_sku=f"{product.original_sku or product.sku}_{idx}",
                category=product.category, country=product.country, language=product.language,
                feed_no=idx, intent=item["intent"].strip(), brand_title=title, title_chars=len(title),
                brand_body_copy=body, body_chars=len(body), evidence_ids=", ".join(ids),
                validation_status=status, review_notes=item.get("review_notes", "").strip(),
            ))
        return feeds


def generator_for(mode: str):
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    model = os.getenv("OPENAI_MODEL", "gpt-5.6-luna").strip()
    if mode == "openai" and not api_key:
        raise RuntimeError("OPENAI_API_KEY가 설정되지 않았습니다.")
    if mode in {"openai", "auto"} and api_key:
        return OpenAIGenerator(api_key, model)
    return RuleBasedGenerator()


def process_product(product: ProductInput, *, max_feeds: int = 4, title_limit: int = 30, body_limit: int = 60, generator_mode: str = "auto") -> ProductResult:
    started = time.perf_counter()
    result = ProductResult(product=product)
    evidence: list[Evidence] = []

    if product.key_features:
        source_type = "PIM/CMS"
        source_record = product.source_record_id or product.sku
        for idx, feature in enumerate(product.key_features, 1):
            evidence.append(Evidence(
                claim_id=f"{product.sku}-C{idx:02d}", sku=product.sku, category=product.category,
                source_type=source_type, source_record_id=source_record, source_section="Key Features",
                source_text=feature, intent_candidate=classify_intent(feature, product.category, product.country), validation_status="Needs Review",
                notes="업로드된 구조화 필드. 데이터 오너/최신성 확인 전.",
            ))
    else:
        if not product.url and country_config(product.country):
            try:
                product.url = LGSitemapResolver.resolve(product.country, product.sku, product.category)
                if product.url:
                    if not product.category:
                        product.category = category_from_url(product.url)
                    result.issues.append(Issue(product.sku, "resolve", "info", "URL_RESOLVED", product.url))
            except Exception as exc:
                result.issues.append(Issue(product.sku, "resolve", "warning", "URL_RESOLVE_FAILED", str(exc)))

    if not evidence and product.url:
        try:
            if not product.category:
                product.category = category_from_url(product.url)
            extractor = PDPExtractor()
            html = extractor.fetch(product.url)
            product.product_name, evidence, extract_issues = extractor.extract(product, html)
            result.issues.extend(extract_issues)
        except Exception as exc:
            result.issues.append(Issue(product.sku, "fetch", "error", "PDP_FETCH_FAILED", str(exc)))
    elif not evidence:
        result.issues.append(Issue(
            product.sku, "input", "error", "SOURCE_REQUIRED",
            f"{product.country} PDP URL을 찾지 못했습니다. 정확한 PDP URL 또는 PIM/CMS Key Features가 필요합니다.",
        ))

    result.evidence = evidence
    if evidence:
        try:
            generator = generator_for(generator_mode)
            result.feeds = generator.generate(product, evidence, max_feeds, title_limit, body_limit)
            if isinstance(generator, RuleBasedGenerator):
                reason = "규칙 기반 모드를 선택해 검토용 초안을 생성했습니다." if generator_mode == "rules" else "API 키가 없어 규칙 기반 검토용 초안을 생성했습니다."
                result.issues.append(Issue(product.sku, "generate", "warning", "RULE_DRAFT", reason))
        except Exception as exc:
            result.issues.append(Issue(product.sku, "generate", "error", "GENERATION_FAILED", str(exc)))

    if not result.feeds and evidence:
        result.issues.append(Issue(product.sku, "generate", "warning", "NO_FEEDS", "생성 가능한 Feed가 없습니다."))
    result.elapsed_seconds = round(time.perf_counter() - started, 2)
    return result


def result_to_dict(result: ProductResult) -> dict[str, Any]:
    return {
        "product": asdict(result.product),
        "feeds": [asdict(x) for x in result.feeds],
        "evidence": [asdict(x) for x in result.evidence],
        "issues": [asdict(x) for x in result.issues],
        "elapsed_seconds": result.elapsed_seconds,
    }
