from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .core import ProductResult


HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")


def _write_table(ws, headers: list[str], rows: Iterable[list[object]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, 1):
        values = [str(ws.cell(r, col_idx).value or "") for r in range(1, min(ws.max_row, 200) + 1)]
        width = min(max(len(header) + 2, *(len(v) + 2 for v in values)), 65)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook(results: list[ProductResult], output_path: Path, settings: dict[str, object]) -> Path:
    """Export the delivery feed schema; review metadata remains in the app, not the feed."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Feeds"
    headers = ["id", "title", "link", "image_link", "description", "price"]
    rows = []
    for result in results:
        for feed in result.feeds:
            rows.append([
                feed.virtual_sku,
                feed.brand_title,
                result.product.url,
                result.product.image_link,
                feed.brand_body_copy,
                result.product.price,
            ])
    _write_table(ws, headers, rows)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 70
    ws.column_dimensions["F"].width = 14
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
